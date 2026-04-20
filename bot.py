import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import datetime
import os
from selenium.webdriver.support.ui import Select


# Ruta al archivo Excel local
# Última fila de Excel procesada con éxito (reanudar si el bot se corta). Borrar el archivo para empezar de cero.
ARCHIVO_PROGRESO = "progreso_facturador.txt"
ARCHIVO_LOG = "facturador_historial.log"


def cargar_ultima_fila_ok():
    try:
        with open(ARCHIVO_PROGRESO, encoding="utf-8") as f:
            return int(f.read().strip())
    except (FileNotFoundError, ValueError, OSError):
        return 0


def guardar_ultima_fila_ok(fila):
    with open(ARCHIVO_PROGRESO, "w", encoding="utf-8") as f:
        f.write(str(int(fila)))


def registrar_log(mensaje):
    linea = f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {mensaje}\n"
    with open(ARCHIVO_LOG, "a", encoding="utf-8") as f:
        f.write(linea)


def limpiar_progreso():
    """Quita el checkpoint cuando el Excel terminó por completo (no aplica al próximo archivo)."""
    try:
        os.remove(ARCHIVO_PROGRESO)
    except OSError:
        pass


def resumen_progreso_excel(archivo_excel, nombre_hoja):
    """
    Devuelve dict con última fila OK, cantidad de filas con código de venta y pendientes,
    o None si no se puede leer.
    """
    if not os.path.isfile(archivo_excel):
        return None
    try:
        wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            return None
        sh = wb[nombre_hoja]
        ultima = cargar_ultima_fila_ok()
        con_codigo = 0
        pendientes = 0
        for r in range(2, sh.max_row + 1):
            v = sh.cell(row=r, column=1).value
            if v is None or (isinstance(v, str) and not str(v).strip()):
                continue
            con_codigo += 1
            if r > ultima:
                pendientes += 1
        wb.close()
        hechas = max(0, con_codigo - pendientes)
        return {
            "ultima_fila_checkpoint": ultima,
            "facturas_en_excel": con_codigo,
            "pendientes": pendientes,
            "estimadas_ya_cargadas": hechas,
        }
    except Exception:
        return None


def ejecutar_facturador(
    archivo_excel,
    cuit,
    password,
    nombre_hoja,
    log_print=print,
    al_terminar_lote=None,
):
    """Ejecuta el bot: ruta al .xlsx, CUIT, clave, nombre de la hoja en el libro."""

    def _log(*args, **kwargs):
        """Evita que flush=True rompa un log personalizado que no acepta kwargs."""
        kwargs.pop("flush", None)
        log_print(*args, **kwargs)

    if not os.path.isfile(archivo_excel):
        raise FileNotFoundError(f"No se encuentra el archivo: {archivo_excel}")

    # Cargar el libro de Excel
    workbook = openpyxl.load_workbook(archivo_excel, data_only=True)
    if nombre_hoja not in workbook.sheetnames:
        raise ValueError(
            f"No existe la hoja '{nombre_hoja}'. Hojas en el archivo: {workbook.sheetnames}"
        )
    sheet = workbook[nombre_hoja]
    hoy = datetime.datetime.now()
    hoy_formateado = hoy.strftime("%d/%m/%Y")
    
    # Obtener el número de filas y columnas
    num_cols = sheet.max_column
    num_rows = 0
    
    
    # Crear listas para almacenar los datos de cada columna
    codigo_venta_list = []
    fecha_list = []
    codigo_servicio_list = []
    servicio_list = []
    total_list = []
    converted_date = []
    data_dict_list = []
    
    # Iterar sobre las filas del Excel (max_row del sheet, no num_rows: aún está en 0)
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        if any(value is not None for value in row):
            codigo_venta_list.append(row[0])
            fecha_list.append(row[1])
            codigo_servicio_list.append(row[2])
            servicio_list.append(row[3])
            total_list.append(row[4])
            num_rows += 1  # cantidad de filas con datos
        else:
            break
    
    
    for i in fecha_list:
        if i is not None:
            if isinstance(i, (datetime.datetime, datetime.date)):
                converted_date.append(i.strftime("%d/%m/%Y"))
            else:
                converted_date.append(str(i))
    
    
    def get_value(list, index, num):
        var = ""
        for index, row in enumerate(range(num)):
            if index < len(list):
                var = list[index]
            return var
    
    
    # Configurar el navegador
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("enable-automation")
    options.add_experimental_option("detach", True)
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-extensions")
    options.add_argument("--dns-prefetch-disable")
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=Service(
        ChromeDriverManager().install()), options=options)
    actions = ActionChains(driver)
    # URL del formulario
    url_formulario = "https://auth.afip.gob.ar/contribuyente_/login.xhtml"
    
    # Abrir el formulario en el navegador
    try:
        # Captura la ventana actual
        driver.get(url_formulario)
        time.sleep(2)  # Esperar a que la página cargue inicialmente
    
        # Esperar y llenar CUIT
        wait = WebDriverWait(driver, 15)
        input_cuit = wait.until(
            EC.presence_of_element_located((By.ID, "F1:username")))
        input_cuit.clear()
        input_cuit.send_keys(cuit)
        time.sleep(1)
    
        # Esperar y hacer click en Siguiente
        next_btn = wait.until(
            EC.element_to_be_clickable((By.ID, "F1:btnSiguiente")))
        next_btn.click()
        time.sleep(3)  # Esperar a que cargue la página de contraseña
    
        # Esperar explícitamente a que el campo de contraseña esté disponible y visible
        input_password = wait.until(
            EC.presence_of_element_located((By.ID, "F1:password"))
        )
        # Esperar adicional para asegurar que el campo esté completamente cargado
        wait.until(EC.element_to_be_clickable((By.ID, "F1:password")))
        time.sleep(1)
    
        # Limpiar y escribir la contraseña
        input_password.clear()
        input_password.send_keys(password)
        time.sleep(1)
    
        # Esperar y hacer click en Ingresar
        next_btn = wait.until(
            EC.element_to_be_clickable((By.ID, "F1:btnIngresar")))
        next_btn.click()
        time.sleep(3)  # Esperar a que cargue el dashboard
    
        """ DASHBOARD """
        wait = WebDriverWait(driver, 10)
        comprobantes_en_linea = wait.until(EC.visibility_of_element_located(
            (By.XPATH, "//*[contains(text(), 'Comprobantes en línea')]")))
        comprobantes_en_linea.click()
    
        time.sleep(2)
        current_window = driver.current_window_handle
        # Cambia el control a la nueva ventana que se abrió
        for window_handle in driver.window_handles:
            if window_handle != current_window:
                driver.switch_to.window(window_handle)
                break
    
        time.sleep(1)
        driver.execute_script("document.querySelector('div#encabezado_logo_afip img').src = 'https://res.cloudinary.com/practicaldev/image/fetch/s--Rr7K5gOm--/c_limit%2Cf_auto%2Cfl_progressive%2Cq_auto%2Cw_880/https://dbalas.gallerycdn.vsassets.io/extensions/dbalas/vscode-html2pug/0.0.2/1532242577062/Microsoft.VisualStudio.Services.Icons.Default'")
        driver.execute_script(
            "document.querySelector('div#encabezado_logo_afip img').style.width = '4rem'")
    
        # Primera vez en RCEL: empresa + Generar comprobantes + punto de venta (mismo orden que antes).
        # Tras "Menú Principal": solo volver a pulsar Generar comprobantes; no repetir empresa.
        # No borrar filas del Excel: el for avanza de fila solo.
        empresa_seleccionada = False
        ultima_fila_ok = cargar_ultima_fila_ok()
        if ultima_fila_ok > 0:
            _log(
                f"Checkpoint: reanudando; filas 2–{ultima_fila_ok} ya facturadas (omitir). "
                f"Borrá '{ARCHIVO_PROGRESO}' para procesar todo desde el inicio.",
                flush=True,
            )
    
        for index, row in enumerate(range(num_rows)):
            current_row = index + 2
            _log("current row: ", current_row)
    
            if current_row <= ultima_fila_ok:
                _log(
                    f"  → Omitida (ya cargada según {ARCHIVO_PROGRESO})",
                    flush=True,
                )
                continue
    
            codigo_venta = sheet.cell(row=current_row, column=1).value
    
            # Saltar filas que no tienen código de venta (ya fueron procesadas como servicios adicionales)
            if codigo_venta is None or codigo_venta == "":
                _log(
                    f"Saltando fila {current_row} - no tiene código de venta (servicio adicional)")
                continue
    
            try:
                # Elegir empresa solo la primera vez que procesamos una fila con datos (no usar index==0: la fila 2 puede estar vacía)
                if not empresa_seleccionada:
                    btn_empresa = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "btn_empresa")))
                    time.sleep(1)
                    btn_empresa.click()
                    time.sleep(1)
                    empresa_seleccionada = True

                # Cada factura: id btn_gen_cmp (mismo que antes)
                gen_cmp = wait.until(EC.element_to_be_clickable((By.ID, "btn_gen_cmp")))
                try:
                    gen_cmp.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", gen_cmp)
                time.sleep(2)

                puntos_de_venta = driver.find_element(
                    By.XPATH, "//select[@name='puntoDeVenta']").click()
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                actions.send_keys(Keys.ENTER)
                actions.perform()
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(2)
                actions.send_keys(Keys.TAB)
                actions.perform()

                actions.send_keys(Keys.TAB)
                actions.perform()

                """ continuar al formulario (mismo flujo que con ENTER sobre el foco) """
                actions.send_keys(Keys.ENTER)
                actions.perform()

                # Obtener valores directamente del Excel para esta fila
                fecha_excel = sheet.cell(row=current_row, column=2).value
                codigo_venta_excel = sheet.cell(row=current_row, column=1).value
                codigo_servicio_excel = sheet.cell(row=current_row, column=3).value
                servicio_excel = sheet.cell(row=current_row, column=4).value
                total_excel = sheet.cell(row=current_row, column=5).value

                # Convertir fecha a formato string si es necesario
                if fecha_excel is not None:
                    if isinstance(fecha_excel, datetime.datetime):
                        fecha_str = fecha_excel.strftime("%d/%m/%Y")
                    else:
                        fecha_str = str(fecha_excel)
                else:
                    fecha_str = ""

                # Esperar el input fechaEmisionComprobante (id=fc), seleccionarlo y tipear la fecha de hoy
                wait = WebDriverWait(driver, 10)
                fecha = wait.until(EC.presence_of_element_located((By.ID, "fc")))
                time.sleep(1)
                fecha.click()  # seleccionar el input para que tenga foco
                fecha.clear()
                fecha.send_keys(hoy_formateado)  # insertar fecha de hoy (dd/mm/yyyy)
                time.sleep(2)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)

                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)

                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)

                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)

                actions.send_keys(fecha_str)
                actions.perform()
                time.sleep(2)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(0.5)

                actions.send_keys(fecha_str)
                actions.perform()
                time.sleep(1)

                # Esperar y hacer click en actividad
                actividad = wait.until(
                    EC.element_to_be_clickable((By.ID, "actiAsociadaId")))
                actividad.click()
                time.sleep(1)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(1)

                # Esperar y llenar referencia
                ref = wait.until(EC.presence_of_element_located(
                    (By.ID, "refComEmisor")))
                ref.clear()
                ref.send_keys(str(codigo_venta_excel) if codigo_venta_excel else "")
                time.sleep(1)

                # Esperar y hacer click en continuar
                continuar = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//input[@value='Continuar >']")))
                continuar.click()
                time.sleep(3)  # Esperar a que cargue la siguiente página

                # Esperar y seleccionar consumidor final
                consumidor_final = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//select[@name='idIVAReceptor']")))
                consumidor_final.click()
                time.sleep(1)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(0.5)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(1)

                # Esperar y seleccionar condiciones de venta
                condiciones_de_venta = wait.until(
                    EC.element_to_be_clickable((By.ID, "formadepago7")))
                condiciones_de_venta.click()
                time.sleep(2)

                # Esperar y hacer click en continuar
                continuar = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//input[@value='Continuar >']")))
                continuar.click()
                time.sleep(3)  # Esperar a que cargue la página de detalles

                # Esperar y llenar código de artículo
                codigo_articulo = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//input[@class='soloTexto']")))
                codigo_articulo.clear()
                codigo_articulo.send_keys(
                    str(codigo_servicio_excel) if codigo_servicio_excel else "")
                time.sleep(1)

                # Esperar y llenar nombre de artículo
                nombre_articulo = wait.until(
                    EC.presence_of_element_located((By.ID, "detalle_descripcion1")))
                nombre_articulo.clear()
                nombre_articulo.send_keys(
                    str(servicio_excel) if servicio_excel else "")
                time.sleep(1)

                # Esperar y seleccionar unidad de medida
                unidad_medida = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//select[@name='detalleMedida']")))
                unidad_medida.click()
                time.sleep(1)
                for _ in range(7):
                    actions.send_keys(Keys.ARROW_DOWN)
                    time.sleep(0.2)

                actions.perform()
                time.sleep(1)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(1)

                # Esperar y llenar precio
                precio = wait.until(EC.presence_of_element_located(
                    (By.ID, "detalle_precio1")))
                precio.clear()
                precio.send_keys(str(total_excel) if total_excel else "")
                time.sleep(2)

                # Limpiar la lista de servicios adicionales para esta factura
                data_dict_list.clear()

                # Manejar servicios adicionales (filas sin código de venta)
                # Verificar múltiples servicios adicionales consecutivos
                next_row = current_row + 1
                max_excel_row = sheet.max_row

                # Loop para agregar todos los servicios adicionales consecutivos
                while next_row <= max_excel_row:
                    next_codigo_venta = sheet.cell(row=next_row, column=1).value

                    # Si hay un nuevo código de venta, salir del loop
                    if next_codigo_venta is not None and next_codigo_venta != "":
                        _log(
                            f"Nuevo código de venta detectado: {next_codigo_venta} - finalizando factura actual")
                        break

                    # Obtener los datos del servicio adicional directamente del Excel
                    next_codigo_servicio = sheet.cell(row=next_row, column=3).value
                    next_servicio = sheet.cell(row=next_row, column=4).value
                    next_total = sheet.cell(row=next_row, column=5).value

                    # Verificar que todos los valores estén presentes
                    if all(value is not None for value in [next_codigo_servicio, next_servicio, next_total]):
                        _log(
                            f"Agregando servicio adicional (fila {next_row}): {next_codigo_servicio} - {next_servicio} - {next_total}")

                        # Esperar y presionar "Agregar línea descripción"
                        agregar_servicio = wait.until(EC.element_to_be_clickable(
                            (By.XPATH, "//input[@value='Agregar línea descripción']")))
                        agregar_servicio.click()
                        # Esperar a que se agregue la nueva línea completamente
                        time.sleep(3)

                        # Buscar el último <tr> dentro de la tabla con id "idoperacion" y luego el <input> dentro de ese <tr>
                        # Esto nos llevará directamente a la nueva línea agregada
                        try:
                            # Buscar todos los <tr> dentro de la tabla con id "idoperacion"
                            tabla_operacion = wait.until(
                                EC.presence_of_element_located((By.ID, "idoperacion")))
                            filas_tabla = tabla_operacion.find_elements(
                                By.XPATH, ".//tbody//tr[.//input[@class='soloTexto']]")

                            if len(filas_tabla) > 0:
                                # Tomar el último <tr> (la nueva línea agregada)
                                ultima_fila = filas_tabla[-1]

                                # Buscar el <input> con class 'soloTexto' dentro del último <tr> <td>
                                nuevo_campo_codigo = ultima_fila.find_element(
                                    By.XPATH, ".//td//input[@class='soloTexto']")

                                # Hacer click directamente en el campo de código de la nueva línea sin hacer scroll
                                nuevo_campo_codigo.click()
                                time.sleep(0.5)

                                # Ingresar código de servicio
                                nuevo_campo_codigo.clear()
                                nuevo_campo_codigo.send_keys(str(next_codigo_servicio))
                                time.sleep(0.5)
                                nuevo_campo_codigo.send_keys(Keys.TAB)
                                time.sleep(0.5)
                            else:
                                raise Exception(
                                    "No se encontraron filas en la tabla idoperacion")
                        except Exception as e:
                            # Fallback: usar el método anterior si no encontramos el campo
                            _log(
                                f"Advertencia: No se pudo encontrar el último tr/td/input ({str(e)}), usando método alternativo")
                            subtotal = wait.until(EC.element_to_be_clickable(
                                (By.ID, "detalle_subtotal21")))
                            subtotal.click()
                            time.sleep(0.5)
                            actions.send_keys(Keys.TAB)
                            actions.perform()
                            time.sleep(0.5)
                            actions.send_keys(Keys.TAB)
                            actions.perform()
                            time.sleep(0.5)
                            actions.send_keys(str(next_codigo_servicio))
                            actions.perform()
                            time.sleep(0.5)
                            actions.send_keys(Keys.TAB)
                            actions.perform()
                            time.sleep(0.5)

                        # Ingresar servicio (descripción)
                        actions.send_keys(str(next_servicio))
                        actions.perform()
                        time.sleep(1)
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(0.5)
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(0.5)

                        # Seleccionar unidad de medida (7 veces ARROW_DOWN)
                        for _ in range(7):
                            actions.send_keys(Keys.ARROW_DOWN)
                            time.sleep(0.2)
                        time.sleep(1)
                        actions.perform()
                        time.sleep(0.5)

                        # Ingresar precio unitario (Total) - usando TAB para llegar al campo
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(0.5)
                        actions.send_keys(str(next_total))
                        actions.perform()
                        time.sleep(2)  # Esperar a que se guarde el precio

                    # Verificar la siguiente fila
                    next_row += 1

                # Después de agregar todos los servicios adicionales, presionar continuar
                _log("Presionando continuar - finalizando factura")
                time.sleep(2)  # Esperar a que se guarden todos los datos

                # Buscar directamente el botón "Continuar" en lugar de usar TAB
                continuar_btn = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//input[@value='Continuar >']")))
                continuar_btn.click()

                # Esperar a que se procese y cargue la página de confirmación (resumen + Otros Tributos)
                time.sleep(4)

                # Paso 1: "Confirmar Datos..." abre el modal (observarOConfirmar); no saltar al modal sin esto
                _log("Confirmar Datos (btngenerar)...")
                btn_confirmar_datos = wait.until(EC.element_to_be_clickable((By.ID, "btngenerar")))
                try:
                    btn_confirmar_datos.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", btn_confirmar_datos)
                time.sleep(2)

                # Paso 2: modal jQuery "Generar Comprobante" → botón Confirmar
                _log("Confirmando generación del comprobante (modal)...")
                confirmar_modal = (
                    "//div[@role='dialog' and contains(@class,'ui-dialog-buttons')]"
                    "[.//span[contains(@class,'ui-dialog-title') "
                    "and contains(normalize-space(.),'Generar Comprobante')]]"
                    "//div[contains(@class,'ui-dialog-buttonset')]"
                    "//button[.//span[normalize-space(.)='Confirmar']]"
                )
                confirmar_fallback = (
                    "(//div[contains(@class,'ui-dialog-buttonset')]"
                    "//button[.//span[normalize-space()='Confirmar']])[last()]"
                )

                def _click_confirmar(xpath):
                    el = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    try:
                        el.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", el)

                try:
                    _click_confirmar(confirmar_modal)
                except Exception:
                    _click_confirmar(confirmar_fallback)
                time.sleep(5)

                # Después de generar el comprobante, volver al menú principal para procesar la siguiente factura
                _log("Comprobante generado. Volviendo al menú principal...")
                try:
                    # Buscar el botón "Menú Principal" para volver al inicio
                    menu_principal = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//input[@value='Menú Principal']")))
                    menu_principal.click()
                    time.sleep(3)  # Esperar a que cargue el menú principal
                except Exception as e:
                    # Si no encuentra el botón, intentar con otro método
                    _log(
                        f"Advertencia: No se encontró 'Menú Principal', intentando otro método: {str(e)}")
                    try:
                        # Intentar volver usando el botón de la empresa
                        btn_empresa = wait.until(EC.element_to_be_clickable(
                            (By.CLASS_NAME, "btn_empresa")))
                        btn_empresa.click()
                        time.sleep(2)
                    except Exception as e2:
                        _log(f"Error al volver al menú: {str(e2)}")

                _log(
                    "[OK]",
                    f"Fila {current_row} · Ref {codigo_venta_excel} · Factura generada correctamente.",
                    flush=True,
                )
                guardar_ultima_fila_ok(current_row)
                registrar_log(f"OK fila_excel={current_row} ref={codigo_venta_excel}")
                _log(
                    "[INFO]",
                    f"Checkpoint guardado (fila {current_row}); al reanudar se omiten filas hasta acá.",
                    flush=True,
                )
                # El loop continuará automáticamente con la siguiente iteración
            except Exception as invoice_err:
                _log("[ERR]", f"Fila {current_row} · Ref {codigo_venta} · {invoice_err}", flush=True)
                registrar_log(f"ERROR fila_excel={current_row} ref={codigo_venta} {invoice_err}")
                continue
    
        # Salida normal del for: el lote llegó al final sin error (si había filas que procesar)
        if num_rows > 0:
            _log(
                "[INFO]",
                "--- Listo: terminó el Excel de este lote. Progreso reiniciado. ---",
                flush=True,
            )
            limpiar_progreso()
            registrar_log("FIN lote — checkpoint borrado (excel completado)")
            if al_terminar_lote:
                al_terminar_lote()
            else:
                _log(
                    "\n*** Carga terminada: se procesó todo el Excel de este lote. ***\n",
                    flush=True,
                )

    except Exception as e:
        _log("[ERR]", f"{str(e)}", flush=True)

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Facturador AFIP (línea de comandos)")
    parser.add_argument("--excel", default="BOT.xlsx", help="Archivo Excel")
    parser.add_argument("--cuit", default="27321522616", help="CUIT sin guiones")
    parser.add_argument("--password", default="FIfi180686", help="Clave fiscal")
    parser.add_argument("--hoja", default="Facturador", help="Nombre de la hoja")
    args = parser.parse_args()
    ejecutar_facturador(
        args.excel,
        args.cuit,
        args.password,
        args.hoja,
    )
